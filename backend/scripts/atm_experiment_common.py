# -*- coding: utf-8 -*-
"""
ATM 实验：Excel/JSON 解析与评估逻辑的共用模块。

供以下脚本引用：
  - extract_atm_xlsx_to_json.py（仅抽取）
  - run_atm_dataset_experiment.py（实验）
"""
from __future__ import annotations

import json
import re
import sys
from collections import defaultdict
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

BACKEND_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = BACKEND_ROOT.parent

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None  # type: ignore
    load_workbook = None  # type: ignore

DIM_LABEL_TO_KEY = {
    "一致性与规范性": "consistency_and_normativity",
    "完整性": "completeness",
    "必要性": "necessity_traceability",
    "必要性（可追溯性）": "necessity_traceability",
    "可修改性": "modifiability",
}

HEADER_ALIASES = {
    "样本编号": ("样本编号", "编号", "样本ID", "id"),
    "需求路径": ("需求路径", "需求文件", "需求", "raw", "requirements", "req"),
    "用例图路径": ("用例图路径", "用例图", "diagram", "diagram_path"),
    "用例描述路径": ("用例描述路径", "用例描述", "descriptions", "desc", "desc_path"),
    "注入缺陷类型": ("注入缺陷类型", "缺陷类型", "类型"),
    "对应维度": ("对应维度", "维度", "质量维度"),
    "预期扣分": ("预期扣分", "预期", "扣分预期"),
    "定位关键词": ("定位关键词", "关键词", "keywords"),
    "评估对象": ("评估对象", "对象", "artefact", "artifact"),
    "注入缺陷数量": ("注入缺陷数量", "缺陷数量", "n_defects"),
    "缺陷明细": ("缺陷明细", "缺陷列表", "defects_json", "defect_detail"),
}


def norm(s: Any) -> str:
    if s is None:
        return ""
    return str(s).strip()


def build_header_map(header_row: List[Any]) -> Dict[str, int]:
    raw = [norm(h) for h in header_row]
    index_by_std: Dict[str, int] = {}
    for std, aliases in HEADER_ALIASES.items():
        for i, cell in enumerate(raw):
            if cell in aliases:
                index_by_std[std] = i
                break
    return index_by_std


def cell_at(row: Tuple[Any, ...], idx_map: Dict[str, int], std: str) -> str:
    i = idx_map.get(std)
    if i is None or i >= len(row):
        return ""
    return norm(row[i])


def row_dict_sparse(header: List[str], row: Tuple[Any, ...]) -> Dict[str, str]:
    idx_map = build_header_map(header)
    out: Dict[str, str] = {}
    for std in HEADER_ALIASES.keys():
        v = cell_at(row, idx_map, std)
        if v:
            out[std] = v
    return out


def row_dict_full(header: List[str], row: Tuple[Any, ...]) -> Dict[str, str]:
    """抽取 JSON 用：每个标准列都输出（空字符串表示未填）。"""
    idx_map = build_header_map(header)
    return {std: cell_at(row, idx_map, std) for std in HEADER_ALIASES.keys()}


def parse_int(s: str) -> Optional[int]:
    if not s:
        return None
    try:
        return int(float(s))
    except ValueError:
        return None


def parse_defect_json(s: str) -> Optional[List[Dict[str, Any]]]:
    if not s:
        return None
    try:
        data = json.loads(s)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass
    return None


def read_excel_sheet(path: Path, sheet_name: Optional[str] = None) -> Tuple[List[str], List[Tuple[Any, ...]], str, List[str]]:
    if load_workbook is None:
        print("请先安装: pip install openpyxl", file=sys.stderr)
        sys.exit(1)
    wb = load_workbook(path, read_only=True, data_only=True)
    names = wb.sheetnames
    if sheet_name:
        if sheet_name not in names:
            wb.close()
            raise ValueError(f"工作表不存在: {sheet_name}，可选: {names}")
        ws = wb[sheet_name]
        used = sheet_name
    else:
        ws = wb[names[0]]
        used = names[0]
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return [], [], used, names
    header = [str(x).strip() if x is not None else "" for x in rows[0]]
    return header, rows[1:], used, names


def group_rows(header: List[str], data_rows: List[Tuple[Any, ...]]) -> List[Dict[str, Any]]:
    parsed = [row_dict_sparse(header, r) for r in data_rows if any(norm(x) for x in r)]
    groups_map: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    order: List[str] = []
    for rd in parsed:
        sid = rd.get("样本编号", "")
        if not sid:
            continue
        if sid not in groups_map:
            order.append(sid)
        groups_map[sid].append(rd)

    out: List[Dict[str, Any]] = []
    for sid in order:
        rows = groups_map[sid]
        base = dict(rows[0])
        defects: List[Dict[str, str]] = []
        n_declared = None
        for r in rows:
            n_declared = parse_int(r.get("注入缺陷数量", "")) or n_declared

        detail_json = rows[0].get("缺陷明细", "")
        parsed_detail = parse_defect_json(detail_json) if detail_json else None

        if parsed_detail:
            for item in parsed_detail:
                if not isinstance(item, dict):
                    continue
                defects.append(
                    {
                        "注入缺陷类型": norm(item.get("注入缺陷类型", item.get("类型", ""))),
                        "对应维度": norm(item.get("对应维度", item.get("维度", ""))),
                        "预期扣分": norm(item.get("预期扣分", item.get("扣分", ""))),
                        "定位关键词": norm(item.get("定位关键词", item.get("关键词", ""))),
                    }
                )
            dn = n_declared if n_declared and n_declared > 0 else len(defects)
            out.append({"样本编号": sid, "paths": base, "defects": defects, "declared_n": dn})
            continue

        for r in rows:
            if r.get("对应维度"):
                defects.append(
                    {
                        "注入缺陷类型": r.get("注入缺陷类型", ""),
                        "对应维度": r.get("对应维度", ""),
                        "预期扣分": r.get("预期扣分", ""),
                        "定位关键词": r.get("定位关键词", ""),
                    }
                )

        if not defects and n_declared and n_declared >= 2:
            raise ValueError(
                f"样本 {sid}: 注入缺陷数量>=2 但未提供「缺陷明细」JSON，"
                f"也未在同一编号下分多行填写各缺陷，请补全。"
            )

        dn = n_declared if n_declared and n_declared > 0 else (len(defects) if defects else None)
        out.append(
            {
                "样本编号": sid,
                "paths": base,
                "defects": defects,
                "declared_n": dn,
            }
        )
    return out


def build_extract_payload(
    excel_path: Path,
    header: List[str],
    data_rows: List[Tuple[Any, ...]],
    sheet_used: str,
    sheet_names: List[str],
    groups: List[Dict[str, Any]],
) -> Dict[str, Any]:
    def json_safe(x: Any) -> Any:
        if x is None:
            return None
        if isinstance(x, (str, int, float, bool)):
            return x
        return str(x)

    rows_raw: List[List[Any]] = []
    for tup in data_rows:
        rows_raw.append([json_safe(x) for x in tup])
    rows_mapped = [row_dict_full(header, r) for r in data_rows if any(norm(x) for x in r)]
    return {
        "version": 1,
        "source": {
            "excel": str(excel_path.resolve()),
            "sheet": sheet_used,
            "all_sheets": sheet_names,
        },
        "headers": header,
        "rows_raw": rows_raw,
        "rows_mapped": rows_mapped,
        "groups": groups,
        "header_map_resolved": build_header_map(header),
    }


def load_groups_from_dataset_json(path: Path) -> Tuple[List[Dict[str, Any]], Dict[str, Any]]:
    with open(path, "r", encoding="utf-8") as f:
        data = json.load(f)
    if not isinstance(data, dict):
        raise ValueError("数据集 JSON 根节点须为对象")
    groups = data.get("groups")
    if not isinstance(groups, list) or not groups:
        raise ValueError("数据集 JSON 缺少非空字段 groups（请先运行 extract_atm_xlsx_to_json.py）")
    return groups, data


def collect_issues_block(block: Dict[str, Any]) -> str:
    if not block or not isinstance(block, dict):
        return ""
    parts: List[str] = []

    def issue_text(x: Any) -> str:
        if isinstance(x, str):
            return x
        if isinstance(x, dict):
            return str(
                x.get("description")
                or x.get("reason")
                or x.get("message")
                or x.get("detail")
                or ""
            )
        return str(x)

    for attr in (block.get("attributes") or {}).values():
        if isinstance(attr, dict):
            for it in attr.get("issues") or []:
                parts.append(issue_text(it))
    for it in block.get("issues") or []:
        parts.append(issue_text(it))
    return " ".join(parts)


def dim_block(metrics: Dict[str, Any], key: str) -> Optional[Dict[str, Any]]:
    if not metrics or not isinstance(metrics, dict):
        return None
    b = metrics.get(key)
    return b if isinstance(b, dict) else None


def dim_scores_and_text(report: Dict[str, Any], dim_key: str) -> Tuple[Optional[float], Optional[float], str]:
    dm = report.get("diagram_metrics") or {}
    descm = report.get("description_metrics") or {}
    d_block = dim_block(dm, dim_key)
    s_block = dim_block(descm, dim_key)
    d_score = float(d_block["overall"]) if d_block and d_block.get("overall") is not None else None
    s_score = float(s_block["overall"]) if s_block and s_block.get("overall") is not None else None
    texts = []
    if d_block:
        texts.append(collect_issues_block(d_block))
    if s_block:
        texts.append(collect_issues_block(s_block))
    return d_score, s_score, " ".join(texts)


def resolve_dim_key(label: str) -> Optional[str]:
    label = norm(label)
    if label in DIM_LABEL_TO_KEY:
        return DIM_LABEL_TO_KEY[label]
    for k, v in DIM_LABEL_TO_KEY.items():
        if label and k in label:
            return v
    return None


def effective_score(
    artefact: str, d_score: Optional[float], s_score: Optional[float]
) -> Tuple[float, bool, bool]:
    art = norm(artefact) or "自动"
    if art in ("用例图", "diagram", "图"):
        s = d_score if d_score is not None else 1.0
        return s, True, False
    if art in ("用例描述", "描述", "description", "desc"):
        s = s_score if s_score is not None else 1.0
        return s, False, True
    candidates = [x for x in (d_score, s_score) if x is not None]
    if not candidates:
        return 1.0, False, False
    return min(candidates), d_score is not None, s_score is not None


def detected_fixed(artefact: str, d_score: Optional[float], s_score: Optional[float]) -> bool:
    art = norm(artefact) or "自动"
    if art in ("用例图", "diagram", "图"):
        return d_score is not None and d_score < 0.999
    if art in ("用例描述", "描述", "description", "desc"):
        return s_score is not None and s_score < 0.999
    return (d_score is not None and d_score < 0.999) or (s_score is not None and s_score < 0.999)


def keywords(defect_type: str, kw_cell: str) -> List[str]:
    kws: List[str] = []
    if norm(kw_cell):
        kws.extend([x.strip() for x in re.split(r"[,，;；\s]+", kw_cell) if x.strip()])
    if norm(defect_type) and defect_type not in kws:
        kws.append(defect_type)
    return kws


def localized(issue_blob: str, kws: List[str]) -> bool:
    if not issue_blob.strip():
        return False
    for k in kws:
        if k and k in issue_blob:
            return True
    return False


def load_json_file(path: Path) -> Any:
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_descriptions(path: Path) -> List[Dict[str, Any]]:
    data = load_json_file(path)
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        return data.get("use_case_descriptions") or data.get("descriptions") or []
    return []


def resolve_path(base: Path, p: str) -> Path:
    p = norm(p).replace("\\", "/")
    if not p:
        raise ValueError("路径为空")
    cand = Path(p)
    if cand.is_file():
        return cand.resolve()
    rel = base / p
    if rel.is_file():
        return rel.resolve()
    raise FileNotFoundError(f"找不到文件: {p}（已尝试数据根目录 {base}）")


def process_defect_row(
    report: Dict[str, Any],
    defect_type: str,
    dim_label: str,
    expected: float,
    kw_cell: str,
    artefact: str,
) -> Dict[str, Any]:
    dim_key = resolve_dim_key(dim_label)
    if not dim_key:
        raise ValueError(f"无法解析维度: {dim_label}")
    d_score, s_score, issue_blob = dim_scores_and_text(report, dim_key)
    s_eff, _, _ = effective_score(artefact, d_score, s_score)
    detected = detected_fixed(artefact, d_score, s_score)
    kws = keywords(defect_type, kw_cell)
    loc = bool(detected and localized(issue_blob, kws))
    try:
        exp = float(expected)
    except (TypeError, ValueError):
        exp = 0.0
    actual = int(round(exp * (1.0 - min(1.0, max(0.0, s_eff)))))
    return {
        "注入缺陷类型": defect_type,
        "对应维度": dim_label,
        "工具检出": "是" if detected else "否",
        "定位准确": "是" if loc else "否",
        "预期扣分": int(round(exp)) if exp == int(exp) else exp,
        "实际扣分": actual,
        "_expected_num": exp,
        "_actual_num": float(actual),
    }
