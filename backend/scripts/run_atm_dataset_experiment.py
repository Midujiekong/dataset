#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ATM 实验数据集批量评估（仅跑实验，不读 xlsx）。

推荐流程：
  1) python scripts/extract_atm_xlsx_to_json.py --excel <数据集.xlsx>
  2) python scripts/run_atm_dataset_experiment.py --dataset-json experiment_output/atm_dataset_extracted_xxx.json

也可一步指定 --excel（内部会先解析再评估，与旧版行为一致）。

指标与列约定见 atm_experiment_common.py 及 extract 脚本说明。

用法：
  cd backend
  pip install -r requirements.txt
  python scripts/extract_atm_xlsx_to_json.py --excel ../ATM用例模型质量评估实验数据集.xlsx
  python scripts/run_atm_dataset_experiment.py --dataset-json experiment_output/atm_dataset_extracted_xxx.json --data-dir ..

  # 兼容：直接读 Excel（不生成中间 JSON）
  python scripts/run_atm_dataset_experiment.py --excel ../ATM用例模型质量评估实验数据集.xlsx --data-dir ..
"""
from __future__ import annotations

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

BACKEND_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = BACKEND_ROOT.parent
sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import Workbook

from atm_experiment_common import (
    BACKEND_ROOT,
    REPO_ROOT,
    group_rows,
    load_descriptions,
    load_groups_from_dataset_json,
    process_defect_row,
    read_excel_sheet,
    resolve_path,
    load_json_file,
)
from src.services.evaluator.evaluation_service import EvaluationService
from src.services.evaluator.requirements_parser import extract_structured_requirements


def _run_one_eval(
    service: EvaluationService,
    raw_text: str,
    diagram: Dict[str, Any],
    descriptions: List[Dict[str, Any]],
    use_llm_extract: bool,
) -> Dict[str, Any]:
    structured = extract_structured_requirements(raw_text, use_llm=use_llm_extract)
    return service.evaluate(
        {
            "use_case_diagram": diagram,
            "use_case_descriptions": descriptions,
            "requirements": structured,
        }
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="ATM 实验：读取抽取后的 JSON 并批量评估")
    parser.add_argument(
        "--dataset-json",
        type=str,
        default=None,
        help="由 extract_atm_xlsx_to_json.py 生成的数据集 JSON（推荐）",
    )
    parser.add_argument(
        "--excel",
        type=str,
        default=None,
        help="若未传 --dataset-json，可直接指定 xlsx（与旧版等价）",
    )
    parser.add_argument(
        "--data-dir",
        type=str,
        default=str(REPO_ROOT),
        help="需求/用例图/JSON 相对路径所相对的根目录",
    )
    parser.add_argument("--sheet", type=str, default=None, help="与 --excel 联用：工作表名")
    parser.add_argument("--out", type=str, default=None, help="输出结果 xlsx")
    parser.add_argument("--llm", action="store_true", help="评估阶段启用 LLM")
    parser.add_argument("--llm-extract", action="store_true", help="需求抽取启用 LLM")
    parser.add_argument("--save-reports", type=str, default=None, help="每条样本报告 JSONL")
    args = parser.parse_args()

    data_root = Path(args.data_dir)
    out_dir = BACKEND_ROOT / "experiment_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = (BACKEND_ROOT / out_path).resolve()
    else:
        out_path = out_dir / f"atm_experiment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    groups: List[Dict[str, Any]] = []
    if args.dataset_json:
        json_path = Path(args.dataset_json)
        if not json_path.is_file():
            print(f"找不到数据集 JSON: {json_path}", file=sys.stderr)
            return 1
        groups, _meta = load_groups_from_dataset_json(json_path)
    elif args.excel:
        excel_path = Path(args.excel)
        if not excel_path.is_file():
            print(f"找不到 Excel: {excel_path}", file=sys.stderr)
            return 1
        header, data_rows, _, _ = read_excel_sheet(excel_path, args.sheet)
        if not header:
            print("Excel 无表头", file=sys.stderr)
            return 1
        try:
            groups = group_rows(header, data_rows)
        except ValueError as e:
            print(f"分组失败: {e}", file=sys.stderr)
            return 1
    else:
        default_json = out_dir / "atm_dataset_extracted_latest.json"
        if default_json.is_file():
            groups, _ = load_groups_from_dataset_json(default_json)
            print(f"未指定输入，已使用: {default_json}", file=sys.stderr)
        else:
            print(
                "请指定 --dataset-json（推荐）或 --excel。\n"
                "先运行: python scripts/extract_atm_xlsx_to_json.py --excel <你的xlsx>",
                file=sys.stderr,
            )
            return 1

    if not groups:
        print("无有效分组数据", file=sys.stderr)
        return 1

    service = EvaluationService(use_llm=args.llm)
    use_llm_extract = args.llm_extract

    table1: List[List[Any]] = []
    table2: List[List[Any]] = []
    jsonl_f = open(args.save_reports, "w", encoding="utf-8") if args.save_reports else None

    total_defect_entries = 0
    detected_entries = 0
    localized_entries = 0
    consistency_pairs: List[Tuple[float, float]] = []

    for g in groups:
        sid = g["样本编号"]
        paths = g["paths"]
        defects: List[Dict[str, str]] = g["defects"]
        try:
            req_path = resolve_path(data_root, paths.get("需求路径", ""))
            dia_path = resolve_path(data_root, paths.get("用例图路径", ""))
            desc_path = resolve_path(data_root, paths.get("用例描述路径", ""))
        except Exception as e:
            print(f"[跳过] {sid}: {e}", file=sys.stderr)
            continue

        raw_text = req_path.read_text(encoding="utf-8")
        diagram = load_json_file(dia_path)
        descriptions = load_descriptions(desc_path)
        artefact = paths.get("评估对象", "")

        try:
            report = _run_one_eval(service, raw_text, diagram, descriptions, use_llm_extract)
        except Exception as e:
            print(f"[评估失败] {sid}: {e}", file=sys.stderr)
            continue

        if jsonl_f:
            jsonl_f.write(json.dumps({"样本编号": sid, "report": report}, ensure_ascii=False) + "\n")

        if not defects:
            print(f"[提示] {sid}: 无缺陷标注行，已跳过制表。", file=sys.stderr)
            continue

        declared = g.get("declared_n")
        is_single = len(defects) == 1 and (declared is None or declared <= 1)

        if is_single:
            d0 = defects[0]
            try:
                exp_f = float(str(d0.get("预期扣分", "0") or "0").strip() or 0)
            except ValueError:
                exp_f = 0.0
            row_stats = process_defect_row(
                report,
                d0.get("注入缺陷类型", ""),
                d0.get("对应维度", ""),
                exp_f,
                d0.get("定位关键词", ""),
                artefact,
            )
            table1.append(
                [
                    sid,
                    row_stats["注入缺陷类型"],
                    row_stats["对应维度"],
                    row_stats["工具检出"],
                    row_stats["定位准确"],
                    row_stats["预期扣分"],
                    row_stats["实际扣分"],
                ]
            )
            total_defect_entries += 1
            if row_stats["工具检出"] == "是":
                detected_entries += 1
            if row_stats["定位准确"] == "是":
                localized_entries += 1
            if exp_f > 0:
                consistency_pairs.append((exp_f, float(row_stats["实际扣分"])))
        else:
            det_yes = 0
            loc_yes = 0
            n_inj = max(len(defects), int(declared) if declared and declared > 0 else 0)
            if not n_inj:
                n_inj = len(defects)
            for d0 in defects:
                try:
                    exp_f = float(str(d0.get("预期扣分", "0") or "0").strip() or 0)
                except ValueError:
                    exp_f = 0.0
                row_stats = process_defect_row(
                    report,
                    d0.get("注入缺陷类型", ""),
                    d0.get("对应维度", ""),
                    exp_f,
                    d0.get("定位关键词", ""),
                    artefact,
                )
                if row_stats["工具检出"] == "是":
                    det_yes += 1
                if row_stats["定位准确"] == "是":
                    loc_yes += 1
                total_defect_entries += 1
                if row_stats["工具检出"] == "是":
                    detected_entries += 1
                if row_stats["定位准确"] == "是":
                    localized_entries += 1
                if exp_f > 0:
                    consistency_pairs.append((exp_f, float(row_stats["实际扣分"])))

            rate_det = (100.0 * det_yes / n_inj) if n_inj else 0.0
            rate_loc = (100.0 * loc_yes / n_inj) if n_inj else 0.0
            table2.append([sid, n_inj, det_yes, f"{rate_det:.1f}%", loc_yes, f"{rate_loc:.1f}%"])

    if jsonl_f:
        jsonl_f.close()

    sample_count = len({r[0] for r in table1} | {r[0] for r in table2})
    if sample_count == 0:
        sample_count = len(groups)

    sum_expected = sum(p[0] for p in consistency_pairs)
    sum_abs_err = sum(abs(p[1] - p[0]) for p in consistency_pairs)
    if sum_expected > 0:
        score_consistency = max(0.0, 100.0 * (1.0 - sum_abs_err / sum_expected))
    else:
        score_consistency = 0.0

    det_rate = (100.0 * detected_entries / total_defect_entries) if total_defect_entries else 0.0
    loc_rate = (100.0 * localized_entries / total_defect_entries) if total_defect_entries else 0.0

    wb_out = Workbook()
    ws1 = wb_out.active
    ws1.title = "表1_单缺陷"
    ws1.append(["样本编号", "注入缺陷类型", "对应维度", "工具检出", "定位准确", "预期扣分", "实际扣分"])
    for row in table1:
        ws1.append(row)
    ws2 = wb_out.create_sheet("表2_复合缺陷")
    ws2.append(["样本编号", "注入缺陷数量", "工具检出数", "检出率", "定位准确数", "定位准确率"])
    for row in table2:
        ws2.append(row)
    ws3 = wb_out.create_sheet("表3_汇总")
    ws3.append(["评估指标", "数值"])
    ws3.append(["实验样本总数", sample_count])
    ws3.append(["总注入缺陷数", total_defect_entries])
    ws3.append(["缺陷检出率", f"{det_rate:.1f}%"])
    ws3.append(["缺陷定位准确率", f"{loc_rate:.1f}%"])
    ws3.append(["评分一致性", f"{score_consistency:.1f}%"])

    wb_out.save(out_path)
    print(f"已写入: {out_path}")
    print(
        f"汇总: 样本(表1+表2去重)≈{sample_count}, 缺陷条目={total_defect_entries}, "
        f"检出率={det_rate:.1f}%, 定位率={loc_rate:.1f}%, 评分一致性={score_consistency:.1f}%"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
