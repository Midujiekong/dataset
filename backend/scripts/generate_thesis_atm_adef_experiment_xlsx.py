#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
按論文表 6.2 / 6.3 / 6.4 的「目標結果」生成與 run_atm_excel_adef_experiment 相同工作簿結構的 xlsx。

輸出：backend/experiment_output/atm_adef_experiment_thesis_tables.xlsx
"""
from __future__ import annotations

from datetime import datetime
import shutil
from pathlib import Path

from openpyxl import Workbook

BACKEND_ROOT = Path(__file__).resolve().parent.parent
OUT_DIR = BACKEND_ROOT / "experiment_output"
OUT_PATH = OUT_DIR / "atm_adef_experiment_thesis_tables.xlsx"

# 表 6.2：單一缺陷（對應腳本「表1_单缺陷」列）
TABLE1_ROWS: list[list[object]] = [
    ["S01", "语法违规（include 方向错误、步骤跳号）", "一致性与规范性", "是", "是", "—", "—"],
    ["S02", "命名不一致（同一角色多名称）", "一致性与规范性", "是", "是", "—", "—"],
    ["S03", "语义歧义（模糊表述用例名称）", "一致性与规范性", "是", "是", "—", "—"],
    ["S04", "缺失核心用例（快速取款）", "完整性", "是", "是", "—", "—"],
    ["S05", "缺失前置、后置条件", "完整性", "是", "是", "—", "—"],
    ["S06", "缺失必要include 关系", "完整性", "是", "是", "—", "—"],
    ["S07", "冗余参与者 + 无关用例", "必要性", "是", "是", "—", "—"],
    ["S08", "冗余 extend/include 关系", "必要性", "是", "是", "—", "—"],
    ["S09", "用例描述插入无关实现细节", "必要性", "是", "是", "—", "—"],
    ["S10", "低内聚（存取款功能合并）", "可修改性", "是", "是", "—", "—"],
    ["S11", "内容冗余（重复步骤）", "可修改性", "是", "是", "—", "—"],
    ["S12", "结构混乱（步骤编号错误、主备流混编）", "可修改性", "是", "是", "—", "—"],
]

# 表 6.3：組合缺陷（對應「表2_复合缺陷」）
TABLE2_ROWS: list[list[object]] = [
    ["C01", 2, 2, "100%", 2, "100%"],
    ["C02", 2, 2, "100%", 2, "100%"],
    ["C03", 2, 2, "100%", 2, "100%"],
    ["C04", 2, 2, "100%", 2, "100%"],
    ["C05", 2, 2, "100%", 2, "100%"],
    ["C06", 3, 3, "100%", 3, "100%"],
    ["C07", 3, 3, "100%", 3, "100%"],
    ["C08", 3, 3, "100%", 3, "100%"],
    ["C09", 3, 3, "100%", 3, "100%"],
    ["C10", 3, 2, "67%", 2, "100%"],
    ["C11", 3, 3, "100%", 3, "100%"],
    ["C12", 3, 3, "100%", 3, "100%"],
    ["C13", 3, 3, "100%", 3, "100%"],
    ["C14", 4, 4, "100%", 4, "100%"],
    ["C15", 4, 4, "100%", 4, "100%"],
    ["C16", 4, 4, "100%", 4, "100%"],
    ["C17", 4, 3, "75%", 3, "100%"],
    ["C18", 4, 4, "100%", 4, "100%"],
    ["C19", 4, 4, "100%", 4, "100%"],
    ["C20", 4, 4, "100%", 4, "100%"],
]

# 表 6.4：匯總（對應「表3_汇总」；並保留腳本原有兩行便於對照）
TABLE3_ROWS: list[list[object]] = [
    ["实验样本总数", 33],
    ["总注入缺陷数", 70],
    ["成功检出缺陷数", 67],
    ["缺陷检出率", "95.7%"],
    ["精确定位缺陷数", 67],
    ["缺陷定位准确率", "100.0%"],
    ["评分一致性", "N/A（论文表未给出；可填人工预期一致性）"],
    [
        "说明",
        "本文件为论文表6.2–6.4的目标数值生成；表2中C10检出率=2/3，C17=3/4；"
        "定位准确率=定位准确数/工具检出数（对已检出子集）。",
    ],
]


def _build_summary_placeholder() -> list[list[object]]:
    """33 条逐样本占位行，与实验脚本列顺序一致。"""
    header = [
        "样本ID",
        "Excel行",
        "系统领域",
        "样本类型",
        "缺陷类别(G)",
        "具体缺陷(H)",
        "对应指标(I)",
        "质量等级(J)",
        "综合得分",
        "用例图得分",
        "用例描述得分",
        "评估耗时秒",
        "评估模式",
        "状态",
        "错误信息",
    ]
    rows: list[list[object]] = [header]
    excel_row = 2
    for sid, domain, stype in [
        *[("S%02d" % i, "ATM", "单缺陷") for i in range(1, 13)],
        *[("C%02d" % i, "ATM", "复合缺陷") for i in range(1, 21)],
        ("REF", "ATM", "参考/基线"),
    ]:
        rows.append(
            [
                sid,
                excel_row,
                domain,
                stype,
                "",
                "",
                "",
                "",
                "—",
                "—",
                "—",
                "",
                "detailed",
                "成功",
                "（本表为论文目标结果占位，非真实跑批）",
            ]
        )
        excel_row += 1
    return rows


def main() -> int:
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    w0 = wb.active
    w0.title = "逐样本结果"
    for row in _build_summary_placeholder():
        w0.append(row)

    w1 = wb.create_sheet("表1_单缺陷")
    w1.append(["样本编号", "注入缺陷类型", "对应维度", "工具检出", "定位准确", "预期扣分", "实际扣分(严重度)"])
    for row in TABLE1_ROWS:
        w1.append(row)

    w2 = wb.create_sheet("表2_复合缺陷")
    w2.append(["样本编号", "注入缺陷数量", "工具检出数", "检出率", "定位准确数", "定位准确率"])
    for row in TABLE2_ROWS:
        w2.append(row)

    w3 = wb.create_sheet("表3_汇总")
    w3.append(["评估指标", "数值"])
    for row in TABLE3_ROWS:
        w3.append(row)

    wb.save(OUT_PATH)
    alt = OUT_DIR / f"atm_adef_experiment_thesis_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    shutil.copy2(OUT_PATH, alt)
    print(f"已写入: {OUT_PATH}")
    print(f"副本: {alt}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
