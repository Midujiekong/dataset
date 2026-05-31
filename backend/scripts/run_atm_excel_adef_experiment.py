#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
ATM 实验：按 Excel 固定列位读取（与数据集表头一致）：

  A 样本ID | B 系统领域 | C 样本类型 | D 需求文本 | E 用例图JSON | F 用例描述JSON
  G 缺陷类别 | H 具体缺陷 | I 对应指标 | J 质量等级

D/E/F 支持文件路径或内联文本/JSON；G/H/I 用于表1/表2 与工具输出对照（I 列可为简称，脚本会归一）。

K 列及以后可选：预期扣分、定位关键词、评估对象、注入缺陷数量、缺陷明细（JSON）等同旧别名。

用法（本脚本固定：单一大模型 + 整合 LLM prompt，不启用多智能体）：
  python scripts/run_atm_excel_adef_experiment.py --excel ../ATM用例模型质量评估实验数据集.xlsx --data-dir ..

  # 仅规则、无 LLM
  python scripts/run_atm_excel_adef_experiment.py --excel ../x.xlsx --data-dir .. --no-llm

  # 使用表中「预期扣分」参与表3一致性
  python scripts/run_atm_excel_adef_experiment.py --excel ../x.xlsx --data-dir .. --manual-expected
"""
from __future__ import annotations

import argparse
import json
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

BACKEND_ROOT = Path(__file__).resolve().parent.parent
REPO_ROOT = BACKEND_ROOT.parent
sys.path.insert(0, str(BACKEND_ROOT))

from openpyxl import Workbook, load_workbook

from atm_experiment_common import (
    BACKEND_ROOT,
    REPO_ROOT,
    norm,
    parse_defect_json,
    parse_int,
    process_defect_row,
)
from src.services.evaluator.evaluation_service import EvaluationService
from src.services.evaluator.requirements_parser import extract_structured_requirements

# 1-based 列（与当前 ATM 数据集一致）
COL_SAMPLE_ID = 1
COL_DOMAIN = 2
COL_SAMPLE_TYPE = 3
COL_REQ = 4
COL_DIAGRAM = 5
COL_DESC = 6
COL_DEFECT_CATEGORY = 7  # G 缺陷类别
COL_DEFECT_DETAIL = 8  # H 具体缺陷
COL_METRIC = 9  # I 对应指标
COL_QUALITY_LEVEL = 10  # J 质量等级
COL_FIRST_OPTIONAL = 11  # K 起：预期扣分等

# K 列起：表头名 -> 标准名（扫描第 1 行）
EXTRA_HEADER_ALIASES: Dict[str, Tuple[str, ...]] = {
    "注入缺陷类型": ("注入缺陷类型", "缺陷类型", "类型"),
    "对应维度": ("对应维度", "维度", "质量维度", "对应指标"),
    "预期扣分": ("预期扣分", "预期", "扣分预期"),
    "定位关键词": ("定位关键词", "关键词", "keywords"),
    "评估对象": ("评估对象", "对象", "artefact", "artifact"),
    "注入缺陷数量": ("注入缺陷数量", "缺陷数量", "n_defects"),
    "缺陷明细": ("缺陷明细", "缺陷列表", "defects_json", "defect_detail"),
}

# I 列常见简写 -> process_defect_row 所需完整维度名（表驱动 + 子属性关键词）
_METRIC_NORMALIZE = [
    ("一致性与规范性", ("一致性与规范性", "一致性", "规范性", "语法", "命名", "语义", "歧义")),
    ("完整性", ("完整性", "完整", "缺失必要", "缺失", "关系缺失", "前后置")),
    (
        "必要性（可追溯性）",
        ("必要性", "必要性（可追溯性）", "可追溯", "冗余元素", "冗余关系", "无关描述", "无关"),
    ),
    ("可修改性", ("可修改性", "内容冗余", "结构混乱", "低内聚", "内聚")),
    ("重要性", ("重要性", "重要", "优先级")),
]


def _normalize_metric_label(raw: str) -> str:
    """将 I 列「对应指标」简写、子属性名、顿号并列等映射为质量特性全称。"""
    s = norm(raw)
    if not s:
        return ""
    # 基準樣本：不參與「對照維度」扣分表
    if "基准" in s or s in ("高质量基准",):
        return ""
    # 可修改性子串（先於泛「冗余」匹配，避免「內容冗余」誤入必要性）
    if any(
        x in s
        for x in ("内容冗余", "结构混乱", "低内聚", "功能内聚", "结构清晰", "可修改性")
    ):
        return "可修改性"
    if "结构" in s and "混乱" in s:
        return "可修改性"
    # 必要性：冗餘/無關/信息相關性等子屬性
    if any(
        x in s
        for x in (
            "用例冗余",
            "参与者冗余",
            "关系冗余",
            "冗余性",
            "冗余",
            "无关",
            "信息相关性",
            "信息相关",
            "必要性",
            "可追溯",
        )
    ):
        return "必要性（可追溯性）"
    if any(x in s for x in ("完整性", "完整", "缺失", "关系缺失", "前后置", "参与者完整性", "用例完整性")):
        return "完整性"
    if any(x in s for x in ("一致性与规范性", "一致性", "规范性", "语法", "命名", "语义", "歧义", "术语")):
        return "一致性与规范性"
    if any(x in s for x in ("重要性", "优先级")):
        return "重要性"
    for full, keys in _METRIC_NORMALIZE:
        for k in keys:
            if not k:
                continue
            if s == k or k in s:
                return full
    return s


def _cell(ws, row: int, col: int) -> Any:
    return ws.cell(row=row, column=col).value


def _bases(data_root: Path, excel_path: Path) -> List[Path]:
    return [data_root.resolve(), excel_path.parent.resolve()]


def _try_read_file(rel_or_abs: str, bases: List[Path]) -> Optional[str]:
    s = norm(rel_or_abs)
    if not s:
        return None
    p = Path(s)
    candidates = [p] if p.is_absolute() else [b / s for b in bases]
    for c in candidates:
        try:
            if c.is_file():
                return c.read_text(encoding="utf-8")
        except OSError:
            continue
    return None


def _load_requirement_cell(val: Any, bases: List[Path]) -> str:
    if val is None:
        return ""
    if isinstance(val, str) and val.strip().startswith("="):
        raise ValueError("需求列为公式，请先「选择性粘贴为值」再跑实验")
    s = norm(val)
    if not s:
        return ""
    text = _try_read_file(s, bases)
    if text is not None:
        return text
    return s


def _load_json_cell(val: Any, bases: List[Path], what: str) -> Any:
    if val is None:
        raise ValueError(f"{what} 为空")
    if isinstance(val, str) and val.strip().startswith("="):
        raise ValueError(f"{what} 列为公式，请先粘贴为值")
    s = norm(val)
    if not s:
        raise ValueError(f"{what} 为空")
    raw = _try_read_file(s, bases)
    if raw is not None:
        s = raw
    try:
        return json.loads(s)
    except json.JSONDecodeError as e:
        raise ValueError(f"{what} 不是合法 JSON：{e}") from e


def _diagram_from_cell(val: Any, bases: List[Path]) -> Dict[str, Any]:
    data = _load_json_cell(val, bases, "用例图(D列)")
    if not isinstance(data, dict):
        raise ValueError("用例图须为 JSON 对象")
    return data


def _descriptions_from_cell(val: Any, bases: List[Path]) -> List[Dict[str, Any]]:
    data = _load_json_cell(val, bases, "用例描述(F列)")
    if isinstance(data, list):
        return data
    if isinstance(data, dict):
        lst = data.get("use_case_descriptions") or data.get("descriptions")
        if isinstance(lst, list):
            return lst
    raise ValueError("用例描述须为 JSON 数组或包含 use_case_descriptions / descriptions 的对象")


def _scan_optional_columns(ws, header_row: int) -> Dict[str, int]:
    """K 列起：按第 1 行表头匹配 预期扣分、缺陷明细 等。"""
    out: Dict[str, int] = {}
    max_col = ws.max_column or COL_FIRST_OPTIONAL
    for c in range(COL_FIRST_OPTIONAL, max_col + 1):
        h = norm(_cell(ws, header_row, c))
        if not h:
            continue
        for std, aliases in EXTRA_HEADER_ALIASES.items():
            if h in aliases:
                out[std] = c
                break
    return out


def _read_extra_row(ws, row: int, col_map: Dict[str, int]) -> Dict[str, str]:
    d: Dict[str, str] = {}
    for std, c in col_map.items():
        v = _cell(ws, row, c)
        if v is None:
            d[std] = ""
        elif isinstance(v, (int, float)):
            d[std] = str(v).strip()
        else:
            d[std] = str(v).strip()
    return d


def _defects_from_extras(
    ex: Dict[str, str], default_expected: float
) -> Tuple[List[Dict[str, str]], Optional[int]]:
    """
    缺陷明细 JSON 优先；否则用 G/H/I（缺陷类别、具体缺陷、对应指标）构造一条；
    再否则用「对应维度」旧列名。
    """
    detail = parse_defect_json(ex.get("缺陷明细", ""))
    defects: List[Dict[str, str]] = []
    n_declared = parse_int(ex.get("注入缺陷数量", ""))

    if detail:
        for item in detail:
            if not isinstance(item, dict):
                continue
            raw_dim = norm(item.get("对应维度", item.get("维度", item.get("对应指标", ""))))
            dim_final = _normalize_metric_label(raw_dim)
            if not dim_final:
                continue
            defects.append(
                {
                    "注入缺陷类型": norm(item.get("注入缺陷类型", item.get("类型", ""))),
                    "对应维度": dim_final,
                    "预期扣分": norm(item.get("预期扣分", item.get("扣分", ""))),
                    "定位关键词": norm(item.get("定位关键词", item.get("关键词", ""))),
                }
            )
        dn = n_declared if n_declared and n_declared > 0 else len(defects)
        return defects, dn

    dim_raw = norm(ex.get("对应维度") or ex.get("对应指标", ""))
    dim = _normalize_metric_label(dim_raw)
    cat = norm(ex.get("注入缺陷类型") or ex.get("缺陷类别", ""))
    spec = norm(ex.get("具体缺陷", ""))
    inj_type = " ".join(x for x in [cat, spec] if x).strip() or cat or spec
    if dim:
        if not inj_type:
            inj_type = "（无缺陷类别/具体缺陷）"
        exp_s = norm(ex.get("预期扣分", ""))
        if not exp_s and default_expected > 0:
            de = float(default_expected)
            exp_s = str(int(de)) if de == int(de) else str(de)
        kw = norm(ex.get("定位关键词", "")) or spec or cat
        defects.append(
            {
                "注入缺陷类型": inj_type,
                "对应维度": dim,
                "预期扣分": exp_s,
                "定位关键词": kw,
            }
        )
        dn = n_declared if n_declared and n_declared > 0 else None
        return defects, dn

    if ex.get("对应维度"):
        exp_s = norm(ex.get("预期扣分", ""))
        if not exp_s and default_expected > 0:
            de = float(default_expected)
            exp_s = str(int(de)) if de == int(de) else str(de)
        raw_d = norm(ex.get("对应维度", ""))
        dim_only = _normalize_metric_label(raw_d)
        if not dim_only:
            return [], n_declared
        defects.append(
            {
                "注入缺陷类型": ex.get("注入缺陷类型", ""),
                "对应维度": dim_only,
                "预期扣分": exp_s,
                "定位关键词": ex.get("定位关键词", ""),
            }
        )
        dn = n_declared if n_declared and n_declared > 0 else None
        return defects, dn

    return [], n_declared


def _run_eval(
    service: EvaluationService,
    raw_text: str,
    diagram: Dict[str, Any],
    descriptions: List[Dict[str, Any]],
    use_llm_extract: bool,
    use_harmonic_overall: bool,
) -> Dict[str, Any]:
    structured = extract_structured_requirements(raw_text, use_llm=use_llm_extract)
    policy = "mean" if not use_harmonic_overall else "harmonic"
    return service.evaluate(
        {
            "use_case_diagram": diagram,
            "use_case_descriptions": descriptions,
            "requirements": structured,
            "evaluation_mode": "quick",
            "overall_score_policy": policy,
            "force_single_llm": True,
        }
    )


def main() -> int:
    parser = argparse.ArgumentParser(description="ATM：A列样本ID + D/E/F 三列输入，逐行实验与三张表")
    parser.add_argument("--excel", type=str, required=True, help="xlsx 路径")
    parser.add_argument("--sheet", type=str, default=None, help="工作表名（默认第一个）")
    parser.add_argument("--data-dir", type=str, default=str(REPO_ROOT), help="相对路径解析根目录")
    parser.add_argument(
        "--header-rows",
        type=int,
        default=1,
        help="顶部作为表头的行数（默认 1，数据从第 2 行开始）",
    )
    parser.add_argument("--out", type=str, default=None, help="输出 xlsx")
    parser.add_argument(
        "--llm",
        action="store_true",
        help="显式声明使用评估 LLM（默认已开启，除非 --no-llm）",
    )
    parser.add_argument(
        "--no-llm",
        action="store_true",
        help="关闭评估 LLM（仅规则）",
    )
    parser.add_argument("--llm-extract", action="store_true", help="需求抽取启用 LLM")
    parser.add_argument(
        "--no-harmonic-overall",
        action="store_true",
        help="综合得分用算术平均（默认：调和平均，单模型下仍可用）",
    )
    parser.add_argument(
        "--manual-expected",
        action="store_true",
        help="使用表中「预期扣分」列做表1/表3一致性；默认改为 100×(1-维度得分) 刻画严重程度，无需人工预期分",
    )
    parser.add_argument("--save-reports", type=str, default=None, help="每条样本 JSONL 完整报告")
    parser.add_argument(
        "--default-expected-deduction",
        type=float,
        default=10.0,
        help="与 --manual-expected 联用：表中无预期扣分时使用的默认值",
    )
    parser.add_argument(
        "--throttle-seconds",
        type=float,
        default=0.0,
        help="每行评估前休眠秒数（默认 0）；遇 API 限流时可设 0.3～0.8",
    )
    args = parser.parse_args()

    # 实验跑批：强制单模型路径（忽略 .env 中 MULTI_AGENT_ENABLED；与 Web 端多智能体无关）
    os.environ["MULTI_AGENT_ENABLED"] = "false"

    use_llm = not bool(args.no_llm)
    if bool(args.llm):
        use_llm = True

    excel_path = Path(args.excel)
    if not excel_path.is_file():
        print(f"找不到文件: {excel_path}", file=sys.stderr)
        return 1

    data_root = Path(args.data_dir)
    bases = _bases(data_root, excel_path)
    out_dir = BACKEND_ROOT / "experiment_output"
    out_dir.mkdir(parents=True, exist_ok=True)
    if args.out:
        out_path = Path(args.out)
        if not out_path.is_absolute():
            out_path = (BACKEND_ROOT / out_path).resolve()
    else:
        out_path = out_dir / f"atm_adef_experiment_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    wb_in = load_workbook(excel_path, read_only=False, data_only=True)
    names = wb_in.sheetnames
    if args.sheet:
        if args.sheet not in names:
            print(f"无此工作表: {args.sheet}，可选: {names}", file=sys.stderr)
            return 1
        ws = wb_in[args.sheet]
    else:
        ws = wb_in[names[0]]

    header_rows = max(0, args.header_rows)
    header_row = header_rows if header_rows >= 1 else 1
    optional_col_map = _scan_optional_columns(ws, header_row) if header_rows >= 1 else {}

    service = EvaluationService(use_llm=use_llm, use_multi_agent=False)
    use_llm_extract = args.llm_extract
    use_harmonic = not args.no_harmonic_overall
    use_manual_expected = args.manual_expected

    table1: List[List[Any]] = []
    table2: List[List[Any]] = []
    summary_rows: List[List[Any]] = []

    total_defect_entries = 0
    detected_entries = 0
    localized_entries = 0
    consistency_pairs: List[Tuple[float, float]] = []

    jsonl_f = open(args.save_reports, "w", encoding="utf-8") if args.save_reports else None

    start_data_row = header_rows + 1
    max_r = ws.max_row or 0

    for r in range(start_data_row, max_r + 1):
        sid = norm(_cell(ws, r, COL_SAMPLE_ID))
        d_val = _cell(ws, r, COL_REQ)
        e_val = _cell(ws, r, COL_DIAGRAM)
        f_val = _cell(ws, r, COL_DESC)

        if not sid and d_val in (None, "") and e_val in (None, "") and f_val in (None, ""):
            continue

        if not sid:
            print(f"[跳过 行{r}] 无样本ID(A列)", file=sys.stderr)
            continue

        domain = norm(_cell(ws, r, COL_DOMAIN))
        stype = norm(_cell(ws, r, COL_SAMPLE_TYPE))

        try:
            req_text = _load_requirement_cell(d_val, bases)
            diagram = _diagram_from_cell(e_val, bases)
            descriptions = _descriptions_from_cell(f_val, bases)
        except ValueError as e:
            print(f"[跳过 {sid} 行{r}] {e}", file=sys.stderr)
            summary_rows.append([sid, r, domain, stype, "", "", "", "", "", "", "", "", "", "失败", str(e)])
            continue

        if not req_text:
            print(f"[跳过 {sid} 行{r}] 需求(D列)为空", file=sys.stderr)
            summary_rows.append([sid, r, domain, stype, "", "", "", "", "", "", "", "", "", "失败", "需求为空"])
            continue

        if use_llm and args.throttle_seconds > 0:
            time.sleep(args.throttle_seconds)

        try:
            report = _run_eval(
                service,
                req_text,
                diagram,
                descriptions,
                use_llm_extract,
                use_harmonic_overall=use_harmonic,
            )
        except Exception as e:
            print(f"[评估失败 {sid} 行{r}] {e}", file=sys.stderr)
            summary_rows.append([sid, r, domain, stype, "", "", "", "", "", "", "", "", "", "失败", str(e)])
            continue

        if jsonl_f:
            jsonl_f.write(json.dumps({"样本ID": sid, "行": r, "report": report}, ensure_ascii=False) + "\n")

        ov = report.get("overall_score")
        dg = (report.get("diagram_metrics") or {}).get("overall_score")
        ds = (report.get("description_metrics") or {}).get("overall_score")
        dur = report.get("evaluation_duration_seconds")

        g_cat = norm(_cell(ws, r, COL_DEFECT_CATEGORY))
        g_spec = norm(_cell(ws, r, COL_DEFECT_DETAIL))
        g_metric = norm(_cell(ws, r, COL_METRIC))
        g_level = norm(_cell(ws, r, COL_QUALITY_LEVEL))

        ex: Dict[str, str] = {
            "缺陷类别": g_cat,
            "具体缺陷": g_spec,
            "对应指标": g_metric,
        }
        ex.update(_read_extra_row(ws, r, optional_col_map))

        defects, declared_n = _defects_from_extras(ex, args.default_expected_deduction)
        artefact = ex.get("评估对象", "")

        summary_rows.append(
            [
                sid,
                r,
                domain,
                stype,
                g_cat,
                g_spec,
                g_metric,
                g_level,
                f"{float(ov) * 100:.2f}%" if ov is not None else "",
                f"{float(dg) * 100:.2f}%" if dg is not None else "",
                f"{float(ds) * 100:.2f}%" if ds is not None else "",
                dur if dur is not None else "",
                report.get("evaluation_mode", ""),
                "成功",
                "",
            ]
        )

        if not defects:
            continue

        is_single = len(defects) == 1 and (declared_n is None or declared_n <= 1)

        if is_single:
            d0 = defects[0]
            try:
                raw_exp = float(norm(d0.get("预期扣分", "0")) or 0)
            except ValueError:
                raw_exp = 0.0
            exp_for_math = raw_exp if use_manual_expected else 100.0
            try:
                row_stats = process_defect_row(
                    report,
                    d0.get("注入缺陷类型", ""),
                    d0.get("对应维度", ""),
                    exp_for_math,
                    d0.get("定位关键词", ""),
                    artefact,
                )
            except ValueError as e:
                print(f"[制表警告 {sid}] {e}", file=sys.stderr)
                continue
            pred_disp = row_stats["预期扣分"] if use_manual_expected else "—"
            table1.append(
                [
                    sid,
                    row_stats["注入缺陷类型"],
                    row_stats["对应维度"],
                    row_stats["工具检出"],
                    row_stats["定位准确"],
                    pred_disp,
                    row_stats["实际扣分"],
                ]
            )
            total_defect_entries += 1
            if row_stats["工具检出"] == "是":
                detected_entries += 1
            if row_stats["定位准确"] == "是":
                localized_entries += 1
            if use_manual_expected and raw_exp > 0:
                consistency_pairs.append((raw_exp, float(row_stats["实际扣分"])))
        else:
            det_yes = 0
            loc_yes = 0
            n_inj = max(len(defects), int(declared_n) if declared_n and declared_n > 0 else 0)
            if not n_inj:
                n_inj = len(defects)
            for d0 in defects:
                try:
                    raw_exp = float(norm(d0.get("预期扣分", "0")) or 0)
                except ValueError:
                    raw_exp = 0.0
                exp_for_math = raw_exp if use_manual_expected else 100.0
                try:
                    row_stats = process_defect_row(
                        report,
                        d0.get("注入缺陷类型", ""),
                        d0.get("对应维度", ""),
                        exp_for_math,
                        d0.get("定位关键词", ""),
                        artefact,
                    )
                except ValueError as e:
                    print(f"[制表警告 {sid}] {e}", file=sys.stderr)
                    continue
                if row_stats["工具检出"] == "是":
                    det_yes += 1
                if row_stats["定位准确"] == "是":
                    loc_yes += 1
                total_defect_entries += 1
                if row_stats["工具检出"] == "是":
                    detected_entries += 1
                if row_stats["定位准确"] == "是":
                    localized_entries += 1
                if use_manual_expected and raw_exp > 0:
                    consistency_pairs.append((raw_exp, float(row_stats["实际扣分"])))
            rate_det = (100.0 * det_yes / n_inj) if n_inj else 0.0
            rate_loc = (100.0 * loc_yes / n_inj) if n_inj else 0.0
            table2.append([sid, n_inj, det_yes, f"{rate_det:.1f}%", loc_yes, f"{rate_loc:.1f}%"])

    wb_in.close()
    if jsonl_f:
        jsonl_f.close()

    sample_evaluated = len({row[0] for row in summary_rows if len(row) > 13 and row[13] == "成功"})
    sample_count = sample_evaluated if sample_evaluated else len({row[0] for row in summary_rows})

    sum_expected = sum(p[0] for p in consistency_pairs)
    sum_abs_err = sum(abs(p[1] - p[0]) for p in consistency_pairs)
    if use_manual_expected and sum_expected > 0:
        score_consistency = max(0.0, 100.0 * (1.0 - sum_abs_err / sum_expected))
    else:
        score_consistency = None
    det_rate = (100.0 * detected_entries / total_defect_entries) if total_defect_entries else 0.0
    loc_rate = (100.0 * localized_entries / total_defect_entries) if total_defect_entries else 0.0

    wb_out = Workbook()
    w0 = wb_out.active
    w0.title = "逐样本结果"
    w0.append(
        [
            "样本ID",
            "Excel行",
            "系统领域",
            "样本类型",
            "缺陷类别(G)",
            "具体缺陷(H)",
            "对应指标(I)",
            "质量等级(J)",
            "综合得分",
            "用例图得分",
            "用例描述得分",
            "评估耗时秒",
            "评估模式",
            "状态",
            "错误信息",
        ]
    )
    for row in summary_rows:
        w0.append(row)

    w1 = wb_out.create_sheet("表1_单缺陷")
    w1.append(["样本编号", "注入缺陷类型", "对应维度", "工具检出", "定位准确", "预期扣分", "实际扣分(严重度)"])
    for row in table1:
        w1.append(row)

    w2 = wb_out.create_sheet("表2_复合缺陷")
    w2.append(["样本编号", "注入缺陷数量", "工具检出数", "检出率", "定位准确数", "定位准确率"])
    for row in table2:
        w2.append(row)

    w3 = wb_out.create_sheet("表3_汇总")
    w3.append(["评估指标", "数值"])
    w3.append(["实验样本总数", sample_count])
    w3.append(["总注入缺陷数", total_defect_entries])
    w3.append(["缺陷检出率", f"{det_rate:.1f}%"])
    w3.append(["缺陷定位准确率", f"{loc_rate:.1f}%"])
    w3.append(["评分一致性", f"{score_consistency:.1f}%" if score_consistency is not None else "N/A（未使用人工预期扣分）"])
    w3.append(
        [
            "说明",
            "本脚本固定：单一大模型 + 整合 LLM prompt（不启用多智能体，与 .env 中 MULTI_AGENT 无关）。"
            "实际扣分列=round(100×(1−维度得分))，无需填预期扣分。",
        ]
    )

    wb_out.save(out_path)
    print(f"已写入: {out_path}")
    print(
        f"成功评估样本约 {sample_count} 条；标注缺陷条目 {total_defect_entries}；"
        f"检出率 {det_rate:.1f}% ；定位率 {loc_rate:.1f}%"
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
